#!/usr/bin/env python3
"""Importe les EAN et libellés depuis planogramme_page_*/ *.xlsx → scripts/pages/."""
from __future__ import annotations

import glob
import json
import os
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installez openpyxl : python -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
PAGES_DIR = SCRIPT_DIR / "pages"
# Dossier parent du projet (Nouveau dossier)
PLANO_BASE = SCRIPT_DIR.parent.parent
CATALOG_FALLBACK = SCRIPT_DIR.parent / "public" / "catalog.json"
FALLBACK_PAGES = {1, 48}


def norm(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def deaccent(text: str) -> str:
    return "".join(
        c
        for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


def pick_sheet(workbook: openpyxl.Workbook):
    for name in workbook.sheetnames:
        nl = deaccent(name.lower())
        if "info" in nl or "controle" in nl or nl.startswith("contr"):
            continue
        if "corrig" in nl or (nl.startswith("page") and "controle" not in nl) or "reference" in nl:
            return workbook[name]
    for name in workbook.sheetnames:
        nl = deaccent(name.lower())
        if "info" not in nl and "controle" not in nl:
            return workbook[name]
    return workbook.active


def parse_xlsx(path: Path, page: int) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = pick_sheet(wb)
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    hdr_i = None
    cols: dict[str, int] = {}

    for i, row in enumerate(rows):
        cells = [norm(c) for c in (row or [])]
        low = [deaccent(c.lower()) for c in cells]
        if not any(x == "ean" for x in low):
            continue
        hdr_i = i
        for j, c in enumerate(low):
            if "etag" in c or "etager" in c:
                cols["shelf"] = j
            elif "position" in c:
                cols["position"] = j
            elif c == "ean":
                cols["ean"] = j
            elif "libell" in c or "libelle" in c:
                cols["name"] = j
        if "ean" not in cols:
            for j, c in enumerate(low):
                if "ean" in c and "ancien" not in c and "photo" not in c and "retenu" not in c:
                    cols["ean"] = j
                    break
        if "shelf" not in cols and "position" in cols and cols["position"] > 0:
            cols["shelf"] = cols["position"] - 1
        break

    if hdr_i is None or "ean" not in cols:
        return []

    def cell(row, key: str) -> str:
        idx = cols.get(key)
        if idx is None or idx >= len(row):
            return ""
        return norm(row[idx])

    out: list[dict] = []
    for row in rows[hdr_i + 1 :]:
        if not row:
            continue
        ean = re.sub(r"\D", "", cell(row, "ean"))
        if len(ean) != 13:
            continue
        try:
            shelf = int(float(cell(row, "shelf") or 0))
            position = int(float(cell(row, "position")))
        except (TypeError, ValueError):
            continue
        name = cell(row, "name")
        if not name:
            continue
        out.append(
            {
                "ean": ean,
                "name": name,
                "page": page,
                "shelf": shelf,
                "position": position,
            }
        )
    return out


def load_fallback_pages() -> dict[int, list[dict]]:
    if not CATALOG_FALLBACK.exists():
        return {}
    data = json.loads(CATALOG_FALLBACK.read_text(encoding="utf-8"))
    by_page: dict[int, list[dict]] = {p: [] for p in FALLBACK_PAGES}
    for item in data:
        page = int(item.get("page", 0))
        if page in FALLBACK_PAGES:
            ean = re.sub(r"\D", "", str(item.get("ean", "")))
            if len(ean) == 13:
                by_page[page].append({**item, "ean": ean})
    return by_page


def main() -> None:
    PAGES_DIR.mkdir(parents=True, exist_ok=True)
    total = 0
    per_page: dict[int, int] = {}

    for page in range(2, 48):
        folder = PLANO_BASE / f"planogramme_page_{page}"
        if not folder.is_dir():
            continue
        files = sorted(folder.glob("*.xlsx"))
        if not files:
            print(f"  page {page}: pas de fichier Excel")
            continue
        items = parse_xlsx(files[0], page)
        out_path = PAGES_DIR / f"page-{page:03d}.json"
        out_path.write_text(
            json.dumps(items, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        per_page[page] = len(items)
        total += len(items)
        print(f"  page {page:2d}: {len(items):3d} references <- {files[0].name}")

    fallback = load_fallback_pages()
    for page, items in fallback.items():
        if per_page.get(page):
            continue
        out_path = PAGES_DIR / f"page-{page:03d}.json"
        out_path.write_text(
            json.dumps(items, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        per_page[page] = len(items)
        total += len(items)
        print(f"  page {page:2d}: {len(items):3d} references (ancien catalogue OCR)")

    print(f"\nImport termine : {total} references dans {PAGES_DIR}")


if __name__ == "__main__":
    main()
